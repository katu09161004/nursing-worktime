# -*- coding: utf-8 -*-
"""
看護業務量集計シート（Excel）を生成する。

紙の「看護業務量集計用紙」の代わりに手入力するためのシート。
中分類をプルダウンで選ぶと、小分類のプルダウンがその中分類のものだけに絞られる
（INDIRECT による連動プルダウン）。勤務時間内／外の分数を入れると、
中分類ごとの合計と総計が自動計算される。

main.py の CATEGORIES を唯一の定義元として読むので、区分を変えたら
このスクリプトを流し直せばシートも追従する。

使い方: python make_xlsx.py [出力先.xlsx]
"""

import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from main import CATEGORIES

OUT = sys.argv[1] if len(sys.argv) > 1 else "看護業務量集計シート.xlsx"

ACCENT = "0F766E"
INK = "0F172A"
MUTED = "64748B"
LINE = "DBE3E8"
BAND = "F1F5F7"
WARN = "B45309"

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def sanitize(name: str) -> str:
    """Excelの名前定義に使える文字列にする。

    必ず sub_ を付ける。区分キーには indirect のように Excel の関数名と衝突する
    ものがあり（INDIRECT）、関数名と同じ名前定義は Excel が拒否・修復してしまう。
    接頭辞を付ければ、今後どんなキーを足しても衝突しない。
    """
    return "sub_" + re.sub(r"[^0-9A-Za-z_]", "_", name)


wb = Workbook()

# ---------------------------------------------------------------------------
# 入力シート
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "入力"

ws["A1"] = "看護業務量集計シート"
ws["A1"].font = Font(bold=True, size=15, color=INK)
ws["A2"] = "中分類を選ぶと小分類が絞り込まれます。勤務時間内／外に分数を入力してください。"
ws["A2"].font = Font(size=10, color=MUTED)

ws["A4"] = "部署名"
ws["C4"] = "担当者（匿名ID）"
ws["E4"] = "勤務帯"
ws["G4"] = "日付"
for c in ("A4", "C4", "E4", "G4"):
    ws[c].font = Font(bold=True, size=10, color=MUTED)
for c in ("B4", "D4", "F4", "H4"):
    ws[c].border = box
    ws[c].fill = PatternFill("solid", fgColor="FFFFFF")

HEAD_ROW = 6
headers = ["大分類", "中分類", "小分類", "勤務時間内(分)", "勤務時間外(分)", "計(分)", "備考"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=HEAD_ROW, column=i, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box

FIRST, LAST = HEAD_ROW + 1, HEAD_ROW + 120  # 入力行

for r in range(FIRST, LAST + 1):
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.border = box
        if r % 2 == 0:
            c.fill = PatternFill("solid", fgColor=BAND)
    # 大分類は中分類から自動で引く
    ws.cell(row=r, column=1).value = (
        f'=IFERROR(IF($B{r}="","",VLOOKUP($B{r},定義!$A:$B,2,FALSE)),"")'
    )
    ws.cell(row=r, column=1).font = Font(size=10, color=MUTED)
    # 計 = 内 + 外
    ws.cell(row=r, column=6).value = (
        f'=IF(COUNT($D{r}:$E{r})=0,"",SUM($D{r}:$E{r}))'
    )
    for col in (4, 5, 6):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="right")

widths = {"A": 22, "B": 16, "C": 26, "D": 14, "E": 14, "F": 10, "G": 30}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = f"A{FIRST}"

# ---------------------------------------------------------------------------
# 定義シート（プルダウンの元データ）
# ---------------------------------------------------------------------------
d = wb.create_sheet("定義")
d["A1"] = "中分類"
d["B1"] = "大分類"
d["C1"] = "名前定義キー"
for c in ("A1", "B1", "C1"):
    d[c].font = Font(bold=True, size=10, color="FFFFFF")
    d[c].fill = PatternFill("solid", fgColor=MUTED)

for i, cat in enumerate(CATEGORIES, start=2):
    d.cell(row=i, column=1, value=cat["label"])
    d.cell(row=i, column=2, value=cat.get("group", ""))
    d.cell(row=i, column=3, value=sanitize(cat["key"]))

# 各中分類の小分類リストを縦に並べ、名前定義を張る
col = 5
for cat in CATEGORIES:
    letter = get_column_letter(col)
    d.cell(row=1, column=col, value=cat["label"]).font = Font(bold=True, size=9, color=MUTED)
    for j, sub in enumerate(cat["subs"], start=2):
        d.cell(row=j, column=col, value=sub["label"])
    ref = f"定義!${letter}$2:${letter}${1 + len(cat['subs'])}"
    wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
        sanitize(cat["key"]), attr_text=ref))
    d.column_dimensions[letter].width = 24
    col += 1

d.column_dimensions["A"].width = 16
d.column_dimensions["B"].width = 22
d.column_dimensions["C"].width = 16
d.sheet_state = "hidden"

# ---- プルダウン ----
# dataValidation の formula1 は先頭に "=" を付けない（付けると Excel が
# プルダウンを出さず、ファイル修復を促すことがある）。
# showErrorMessage=True にしないと、設定したエラーメッセージが表示されない。

# 中分類: 定義シートのA列
dv_cat = DataValidation(
    type="list", formula1=f"定義!$A$2:$A${1 + len(CATEGORIES)}",
    allow_blank=True, showErrorMessage=True,
)
dv_cat.error = "一覧から中分類を選んでください"
dv_cat.errorTitle = "入力エラー"
ws.add_data_validation(dv_cat)
dv_cat.add(f"B{FIRST}:B{LAST}")

# 小分類: 選んだ中分類の名前定義を INDIRECT で引く＝連動プルダウン。
# $B7 は行だけ相対にしてあるので、各行が自分の中分類を見る。
dv_sub = DataValidation(
    type="list",
    formula1=f"INDIRECT(VLOOKUP($B{FIRST},定義!$A:$C,3,FALSE))",
    allow_blank=True, showErrorMessage=True,
)
dv_sub.error = "先に中分類を選んでください"
dv_sub.errorTitle = "入力エラー"
ws.add_data_validation(dv_sub)
dv_sub.add(f"C{FIRST}:C{LAST}")

# 分数は0以上の数値のみ
dv_num = DataValidation(
    type="decimal", operator="greaterThanOrEqual", formula1="0",
    allow_blank=True, showErrorMessage=True,
)
dv_num.error = "0以上の数値（分）を入力してください"
dv_num.errorTitle = "入力エラー"
ws.add_data_validation(dv_num)
dv_num.add(f"D{FIRST}:E{LAST}")

# ---------------------------------------------------------------------------
# 集計シート（紙の用紙と同じ形）
# ---------------------------------------------------------------------------
s = wb.create_sheet("集計")
s["A1"] = "集計結果（自動計算）"
s["A1"].font = Font(bold=True, size=15, color=INK)
s["A2"] = "「入力」シートに書くと自動で反映されます。用紙の合計行と同じ並びです。"
s["A2"].font = Font(size=10, color=MUTED)

row = 4
for i, h in enumerate(["大分類", "中分類", "勤務時間内(分)", "勤務時間外(分)", "合計(分)", "割合"], start=1):
    c = s.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="center")
    c.border = box

first_cat_row = row + 1
r = first_cat_row
last_group = None
for cat in CATEGORIES:
    g = cat.get("group", "")
    s.cell(row=r, column=1, value=g if g != last_group else "")
    last_group = g
    s.cell(row=r, column=2, value=cat["label"])
    rng = f"入力!$B${FIRST}:$B${LAST}"
    s.cell(row=r, column=3, value=f'=SUMIF({rng},$B{r},入力!$D${FIRST}:$D${LAST})')
    s.cell(row=r, column=4, value=f'=SUMIF({rng},$B{r},入力!$E${FIRST}:$E${LAST})')
    s.cell(row=r, column=5, value=f"=SUM($C{r}:$D{r})")
    for col in range(1, 7):
        s.cell(row=r, column=col).border = box
    r += 1
last_cat_row = r - 1

# 総計
s.cell(row=r, column=2, value="総計").font = Font(bold=True, color=INK)
s.cell(row=r, column=3, value=f"=SUM($C{first_cat_row}:$C{last_cat_row})").font = Font(bold=True)
s.cell(row=r, column=4, value=f"=SUM($D{first_cat_row}:$D{last_cat_row})").font = Font(bold=True)
s.cell(row=r, column=5, value=f"=SUM($E{first_cat_row}:$E{last_cat_row})").font = Font(bold=True)
for col in range(1, 7):
    s.cell(row=r, column=col).border = box
    s.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BAND)
total_row = r

# 割合（総計が0のときは空欄にする＝0除算を出さない）
for rr in range(first_cat_row, last_cat_row + 1):
    s.cell(row=rr, column=6, value=f'=IF($E${total_row}=0,"",$E{rr}/$E${total_row})')
    s.cell(row=rr, column=6).number_format = "0.0%"

# 大分類（A〜D）別の小計
r = total_row + 2
s.cell(row=r, column=1, value="大分類別").font = Font(bold=True, size=11, color=INK)
r += 1
for i, h in enumerate(["大分類", "", "勤務時間内(分)", "勤務時間外(分)", "合計(分)", "割合"], start=1):
    c = s.cell(row=r, column=i, value=h)
    if h:
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=MUTED)
        c.alignment = Alignment(horizontal="center")
        c.border = box
r += 1
groups = []
for cat in CATEGORIES:
    if cat.get("group") and cat["group"] not in groups:
        groups.append(cat["group"])
first_g_row = r
for g in groups:
    s.cell(row=r, column=1, value=g)
    rng = f"$A${first_cat_row}:$A${last_cat_row}"
    # 大分類セルは先頭行にしか入れていないので、中分類→大分類の対応は定義シートから引く
    s.cell(row=r, column=3, value=f'=SUMPRODUCT((定義!$B$2:$B${1+len(CATEGORIES)}=$A{r})*$C${first_cat_row}:$C${last_cat_row})')
    s.cell(row=r, column=4, value=f'=SUMPRODUCT((定義!$B$2:$B${1+len(CATEGORIES)}=$A{r})*$D${first_cat_row}:$D${last_cat_row})')
    s.cell(row=r, column=5, value=f"=SUM($C{r}:$D{r})")
    s.cell(row=r, column=6, value=f'=IF($E${total_row}=0,"",$E{r}/$E${total_row})')
    s.cell(row=r, column=6).number_format = "0.0%"
    for col in range(1, 7):
        s.cell(row=r, column=col).border = box
    r += 1

# AI効果測定ベースライン
r += 1
s.cell(row=r, column=1, value="★ AI効果測定ベースライン").font = Font(bold=True, size=11, color=WARN)
r += 1
for i, h in enumerate(["AIツール", "対象業務", "勤務時間内(分)", "勤務時間外(分)", "合計(分)", ""], start=1):
    c = s.cell(row=r, column=i, value=h)
    if h:
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=WARN)
        c.alignment = Alignment(horizontal="center")
        c.border = box
r += 1
for cat in CATEGORIES:
    for sub in cat["subs"]:
        if not sub.get("ai_tool"):
            continue
        s.cell(row=r, column=1, value=sub["ai_tool"])
        s.cell(row=r, column=2, value=sub["label"])
        rng = f"入力!$C${FIRST}:$C${LAST}"
        s.cell(row=r, column=3, value=f'=SUMIF({rng},$B{r},入力!$D${FIRST}:$D${LAST})')
        s.cell(row=r, column=4, value=f'=SUMIF({rng},$B{r},入力!$E${FIRST}:$E${LAST})')
        s.cell(row=r, column=5, value=f"=SUM($C{r}:$D{r})")
        for col in range(1, 6):
            s.cell(row=r, column=col).border = box
        r += 1

for col, w in {"A": 24, "B": 22, "C": 16, "D": 16, "E": 12, "F": 10}.items():
    s.column_dimensions[col].width = w

wb.save(OUT)
print(f"作成: {OUT}")
print(f"  中分類 {len(CATEGORIES)} / 小分類 {sum(len(c['subs']) for c in CATEGORIES)}")
